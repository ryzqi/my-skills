from __future__ import annotations

import json
import os
import subprocess
import sys
from pathlib import Path


def _pkg_version(name: str) -> str | None:
    try:
        from importlib.metadata import PackageNotFoundError, version

        try:
            return version(name)
        except PackageNotFoundError:
            return None
    except Exception:
        return None


def _run(cmd: list[str], timeout: int = 30) -> dict:
    try:
        p = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        return {
            "ok": p.returncode == 0,
            "returncode": p.returncode,
            "stdout": (p.stdout or "").strip(),
            "stderr": (p.stderr or "").strip(),
        }
    except FileNotFoundError as e:
        return {"ok": False, "error": f"FileNotFoundError: {e}"}
    except subprocess.TimeoutExpired as e:
        return {"ok": False, "error": f"TimeoutExpired: {e}"}


def _find_soffice_windows() -> list[str]:
    return [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]


def _check_excel_com() -> dict:
    if os.name != "nt":
        return {"supported": False, "reason": "仅 Windows 支持 COM 自动化"}
    try:
        import win32com.client  # type: ignore

        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.Quit()
        return {"supported": True, "ok": True}
    except Exception as e:
        return {"supported": True, "ok": False, "error": str(e)}


def _check_excel_com_open_and_calculate(xlsx_path: Path) -> dict:
    if os.name != "nt":
        return {"supported": False, "reason": "仅 Windows 支持 COM 自动化"}
    try:
        import win32com.client  # type: ignore

        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = None
        try:
            wb = excel.Workbooks.Open(str(xlsx_path))
            excel.CalculateFull()
            wb.Save()
        finally:
            if wb is not None:
                wb.Close(SaveChanges=True)
            excel.Quit()
        return {"supported": True, "ok": True}
    except Exception as e:
        return {"supported": True, "ok": False, "error": str(e)}


def _xlwings_checks(xlsx_path: Path) -> dict:
    try:
        import xlwings as xw

        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        try:
            book = app.books.open(str(xlsx_path))
            book.app.calculate()
            book.save()
            book.close()
        finally:
            app.quit()
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def _openpyxl_checks(xlsx_path: Path) -> dict:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path, data_only=False)
        sheets = wb.sheetnames
        charts = {}
        for name in sheets:
            ws = wb[name]
            charts[name] = len(getattr(ws, "_charts", []))

        formula_count = 0
        for name in sheets:
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        formula_count += 1
        wb.close()

        return {
            "ok": True,
            "sheetnames": sheets,
            "charts_per_sheet": charts,
            "formula_count": formula_count,
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


def _pandas_checks(xlsx_path: Path, out_path: Path) -> dict:
    try:
        import pandas as pd

        df = pd.read_excel(xlsx_path, sheet_name="Data")
        all_sheets = pd.read_excel(xlsx_path, sheet_name=None)
        df.to_excel(out_path, index=False)
        return {
            "ok": True,
            "data_shape": [int(df.shape[0]), int(df.shape[1])],
            "sheets_read": sorted(list(all_sheets.keys())),
            "output_written": str(out_path),
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


def _xlsxwriter_checks(out_path: Path) -> dict:
    try:
        import xlsxwriter

        wb = xlsxwriter.Workbook(out_path)
        ws = wb.add_worksheet("Sheet1")
        ws.write("A1", "a")
        ws.write("B1", "b")
        ws.write_formula("C1", "=A1&B1")
        wb.close()
        return {"ok": True, "output_written": str(out_path)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def _matplotlib_checks(out_path: Path) -> dict:
    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt

        plt.figure(figsize=(4, 2.5))
        plt.plot([1, 2, 3], [1, 4, 2])
        plt.title("示例折线图")
        plt.tight_layout()
        plt.savefig(out_path, dpi=140)
        plt.close()
        return {"ok": True, "output_written": str(out_path)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def _run_recalc(recalc_py: Path, xlsx_path: Path, timeout_seconds: int = 30) -> dict:
    # 优先同进程调用，避免某些 Python 发行版在 subprocess.CreateProcess 上被策略拦截
    try:
        import importlib.util

        spec = importlib.util.spec_from_file_location("_xlsx_recalc", recalc_py)
        if spec is None or spec.loader is None:
            raise RuntimeError("无法加载 recalc.py 的模块规范")
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)  # type: ignore[union-attr]
        if not hasattr(mod, "recalc"):
            raise RuntimeError("recalc.py 未导出 recalc()")
        result = mod.recalc(str(xlsx_path), timeout_seconds)  # type: ignore[attr-defined]
        return {"ok": True, "mode": "inprocess", "result": result}
    except Exception as e:
        inprocess_err = str(e)

    # 回退到子进程执行（某些环境可能更稳定）
    cmd = [sys.executable, str(recalc_py), str(xlsx_path), str(timeout_seconds)]
    r = _run(cmd, timeout=timeout_seconds + 15)
    if not r.get("ok"):
        return {"ok": False, "mode": "subprocess", "inprocess_error": inprocess_err, "run": r}
    try:
        return {
            "ok": True,
            "mode": "subprocess",
            "inprocess_error": inprocess_err,
            "result": json.loads(r.get("stdout", "{}") or "{}"),
            "run": r,
        }
    except Exception as e:
        return {
            "ok": False,
            "mode": "subprocess",
            "inprocess_error": inprocess_err,
            "error": f"无法解析 recalc.py 输出为 JSON: {e}",
            "run": r,
        }


def main() -> None:
    if (
        Path(sys.executable).resolve().as_posix().lower()
        != Path(r"D:\python\python.exe").resolve().as_posix().lower()
    ):
        raise SystemExit("请使用 D:\\python\\python.exe 运行自检（已按你的要求禁用 LibreOffice Python）。")
    root = Path(__file__).resolve().parents[1]
    test_xlsx = root / "test.xlsx"
    recalc_py = root / "xlsx" / "recalc.py"

    report: dict = {
        "python": {
            "executable": sys.executable,
            "version": sys.version,
        },
        "paths": {
            "cwd": str(Path.cwd()),
            "skills_root": str(root),
            "test_xlsx": str(test_xlsx),
            "recalc_py": str(recalc_py),
        },
        "packages": {
            "openpyxl": _pkg_version("openpyxl"),
            "pandas": _pkg_version("pandas"),
            "numpy": _pkg_version("numpy"),
            "matplotlib": _pkg_version("matplotlib"),
            "xlsxwriter": _pkg_version("xlsxwriter"),
            "xlwings": _pkg_version("xlwings"),
            "pywin32": _pkg_version("pywin32"),
        },
        "dependencies": {
            "soffice_candidates": _find_soffice_windows() if os.name == "nt" else [],
        },
        "checks": {},
        "artifacts": [],
    }

    report["checks"]["excel_com"] = _check_excel_com()
    report["checks"]["excel_com_open_and_calculate"] = _check_excel_com_open_and_calculate(test_xlsx)

    soffice_info = []
    for p in report["dependencies"]["soffice_candidates"]:
        soffice_info.append({"path": p, "exists": Path(p).exists()})
    report["checks"]["soffice_paths"] = soffice_info

    if not test_xlsx.exists():
        report["checks"]["test_xlsx_exists"] = {"ok": False, "error": "未找到 test.xlsx"}
        print(json.dumps(report, ensure_ascii=False, indent=2))
        sys.exit(2)
    report["checks"]["test_xlsx_exists"] = {"ok": True}

    report["checks"]["openpyxl"] = _openpyxl_checks(test_xlsx)
    pandas_out = root / "pandas_output.xlsx"
    report["checks"]["pandas"] = _pandas_checks(test_xlsx, pandas_out)
    if report["checks"]["pandas"].get("ok"):
        report["artifacts"].append(str(pandas_out.resolve()))

    xlsxwriter_out = root / "xlsxwriter_output.xlsx"
    report["checks"]["xlsxwriter"] = _xlsxwriter_checks(xlsxwriter_out)
    if report["checks"]["xlsxwriter"].get("ok"):
        report["artifacts"].append(str(xlsxwriter_out.resolve()))

    report["checks"]["xlwings"] = _xlwings_checks(test_xlsx)

    matplotlib_out = root / "matplotlib_output.png"
    report["checks"]["matplotlib"] = _matplotlib_checks(matplotlib_out)
    if report["checks"]["matplotlib"].get("ok"):
        report["artifacts"].append(str(matplotlib_out.resolve()))

    report["checks"]["recalc_test_xlsx"] = _run_recalc(recalc_py, test_xlsx, timeout_seconds=30)

    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
