#!/usr/bin/env python3
"""
Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using LibreOffice
"""

import json
import sys
import subprocess
import os
import platform
from pathlib import Path
try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    load_workbook = None


def setup_libreoffice_macro():
    """Setup LibreOffice macro for recalculation if not already configured"""
    if platform.system() == 'Windows':
        macro_dir = os.path.expandvars(r'%APPDATA%\LibreOffice\4\user\basic\Standard')
    elif platform.system() == 'Darwin':
        macro_dir = os.path.expanduser('~/Library/Application Support/LibreOffice/4/user/basic/Standard')
    else:
        macro_dir = os.path.expanduser('~/.config/libreoffice/4/user/basic/Standard')
    
    macro_file = os.path.join(macro_dir, 'Module1.xba')
    
    if os.path.exists(macro_file):
        with open(macro_file, 'r', encoding='utf-8', errors='ignore') as f:
            if 'RecalculateAndSave' in f.read():
                return True
    
    if not os.path.exists(macro_dir):
        os.makedirs(macro_dir, exist_ok=True)
        try:
            soffice_cmd = find_soffice()
            subprocess.run(
                [soffice_cmd, '--headless', '--terminate_after_init'],
                capture_output=True,
                timeout=10,
            )
        except Exception:
            pass
    
    macro_content = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>'''
    
    try:
        with open(macro_file, 'w', encoding='utf-8', newline='\n') as f:
            f.write(macro_content)
        return True
    except Exception:
        return False


def find_soffice():
    """查找 LibreOffice soffice 可执行文件路径"""
    if platform.system() == 'Windows':
        # Windows 常见安装路径
        common_paths = [
            r'C:\Program Files\LibreOffice\program\soffice.exe',
            r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',
        ]
        for path in common_paths:
            if os.path.exists(path):
                return path
    return 'soffice'  # 其他系统或已在 PATH 中


def recalc(filename, timeout=30):
    """
    Recalculate formulas in Excel file and report any errors

    Args:
        filename: Path to Excel file
        timeout: Maximum time to wait for recalculation (seconds)

    Returns:
        dict with error locations and counts
    """
    if not Path(filename).exists():
        return {'error': f'File {filename} does not exist'}

    abs_path = str(Path(filename).absolute())

    if not setup_libreoffice_macro():
        return {'error': 'Failed to setup LibreOffice macro'}

    soffice_cmd = find_soffice()

    cmd = [
        soffice_cmd, '--headless', '--norestore',
        'vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application',
        abs_path
    ]

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    except FileNotFoundError:
        return {'error': 'LibreOffice not found. Please install LibreOffice from https://www.libreoffice.org/'}
    except subprocess.TimeoutExpired:
        return {'error': f'Recalculation timed out after {timeout} seconds'}

    if result.returncode != 0:
        error_msg = result.stderr or 'Unknown error during recalculation'
        if 'Module1' in error_msg or 'RecalculateAndSave' not in error_msg:
            return {'error': 'LibreOffice macro not configured properly'}
        else:
            return {'error': error_msg}
    
    # Check for Excel errors in the recalculated file - scan ALL cells
    try:
        if load_workbook is None:
            return {
                'error': '缺少依赖 openpyxl，无法读取并扫描工作簿错误；请使用包含 openpyxl 的 Python 运行，或执行：pip install openpyxl'
            }
        wb = load_workbook(filename, data_only=True)
        
        excel_errors = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']
        error_details = {err: [] for err in excel_errors}
        total_errors = 0
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Check ALL rows and columns - no limits
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break
        
        wb.close()
        
        # Build result summary
        result = {
            'status': 'success' if total_errors == 0 else 'errors_found',
            'total_errors': total_errors,
            'error_summary': {}
        }
        
        # Add non-empty error categories
        for err_type, locations in error_details.items():
            if locations:
                result['error_summary'][err_type] = {
                    'count': len(locations),
                    'locations': locations[:20]  # Show up to 20 locations
                }
        
        # Add formula count for context - also check ALL cells
        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
        wb_formulas.close()
        
        result['total_formulas'] = formula_count
        
        return result
        
    except Exception as e:
        return {'error': str(e)}


def main():
    if Path(sys.executable).resolve().as_posix().lower() != Path(r'D:\python\python.exe').resolve().as_posix().lower():
        print("请使用 D:\\python\\python.exe 运行此脚本（已按你的要求禁用 LibreOffice Python）。")
        sys.exit(2)
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        print("\nRecalculates all formulas in an Excel file using LibreOffice")
        print("\nReturns JSON with error details:")
        print("  - status: 'success' or 'errors_found'")
        print("  - total_errors: Total number of Excel errors found")
        print("  - total_formulas: Number of formulas in the file")
        print("  - error_summary: Breakdown by error type with locations")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        sys.exit(1)
    
    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    
    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == '__main__':
    main()
