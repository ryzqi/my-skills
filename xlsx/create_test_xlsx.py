from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


def create_test_xlsx(output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    headers = ["日期", "销量", "单价", "收入"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="FFEFEFEF")
        cell.alignment = Alignment(horizontal="center")

    rows = [
        (date(2025, 1, 1), 10, 9.9),
        (date(2025, 1, 2), 0, 9.9),
        (date(2025, 1, 3), 7, 10.5),
        (date(2025, 1, 4), 12, 8.8),
    ]
    for i, (d, qty, price) in enumerate(rows, start=2):
        ws.cell(i, 1, d)
        ws.cell(i, 2, qty)
        ws.cell(i, 3, price)
        ws.cell(i, 4, f"=B{i}*C{i}")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    for i in range(2, 2 + len(rows)):
        ws.cell(i, 3).number_format = "0.00"
        ws.cell(i, 4).number_format = "0.00"

    chart = BarChart()
    chart.title = "收入柱状图"
    chart.y_axis.title = "收入"
    chart.x_axis.title = "日期"
    data = Reference(ws, min_col=4, min_row=1, max_row=1 + len(rows))
    cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(rows))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "F2")

    calc = wb.create_sheet("Calc")
    calc["A1"] = "总收入"
    calc["B1"] = "=SUM(Data!D2:D5)"
    calc["A2"] = "平均单价"
    calc["B2"] = "=AVERAGE(Data!C2:C5)"
    calc["A3"] = "销量合计"
    calc["B3"] = "=SUM(Data!B2:B5)"
    calc["A4"] = "平均客单价(防除0)"
    calc["B4"] = "=IF(B3=0,0,B1/B3)"

    link = wb.create_sheet("Link")
    link["A1"] = "来自 Calc 的总收入"
    link["B1"] = "=Calc!B1"

    wb.save(output_path)


def main() -> None:
    if Path(sys.executable).resolve().as_posix().lower() != Path(r"D:\python\python.exe").resolve().as_posix().lower():
        raise SystemExit("请使用 D:\\python\\python.exe 运行此脚本（已按你的要求禁用 LibreOffice Python）。")
    output_path = Path("test.xlsx")
    create_test_xlsx(output_path)
    print(f"已生成：{output_path.resolve()}")


if __name__ == "__main__":
    main()
